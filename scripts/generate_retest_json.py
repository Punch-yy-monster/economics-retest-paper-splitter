#!/usr/bin/env python3
"""Generate standardized four-channel retest outputs from a paper text or PDF."""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pypdf import PdfReader


STANDARD_TEXT_PROMPT = "请补充摘要或正文内容，以便继续拆解。"
STANDARD_LANGUAGE_PROMPT = "请确认这是一篇中文文献、英文文献，还是中英混合文献？"

EXIT_INFO_INSUFFICIENT = 1
EXIT_LANGUAGE_UNKNOWN = 2
EXIT_PDF_EXTRACTION_FAILED = 3
EXIT_SCHEMA_INVALID = 4

MIN_TEXT_LENGTH = 180
MIN_PDF_TEXT_LENGTH = 500
MIN_NON_EMPTY_PAGE_RATIO = 0.2

ROOT = Path(__file__).resolve().parents[1]
FULL_SCHEMA_PATH = ROOT / "references" / "output-schema.json"
GENERAL_SCHEMA_PATH = ROOT / "references" / "general-interview-output-schema.json"
ENGLISH_INTERVIEW_SCHEMA_PATH = ROOT / "references" / "english-interview-output-schema.json"
PROFESSIONAL_WRITTEN_SCHEMA_PATH = ROOT / "references" / "professional-written-output-schema.json"
ENGLISH_WRITTEN_SCHEMA_PATH = ROOT / "references" / "english-written-output-schema.json"

ALLOWED_LANGUAGES = {"中文文献", "英文文献", "中英混合文献"}

OUTPUT_VERSION = "2.0.0"
FRAMEWORK_NAME = "five-mandatory-modules-four-channels"

MANDATORY_MODULES = [
    "research_background",
    "research_question",
    "core_conclusion",
    "mechanism_analysis",
    "policy_implication",
]
MODULE_LABELS = {
    "research_background": "研究背景",
    "research_question": "研究问题",
    "core_conclusion": "核心结论",
    "mechanism_analysis": "机制分析",
    "policy_implication": "政策启示",
}
MODULE_QUESTIONS_CN = {
    "research_background": "请用口头方式概括这篇论文的研究背景。",
    "research_question": "这篇论文的核心研究问题是什么？",
    "core_conclusion": "这篇论文最重要的结论是什么？",
    "mechanism_analysis": "作者认为这一结果是通过什么机制产生的？",
    "policy_implication": "这篇论文对现实政策有什么启示？",
}
MODULE_QUESTIONS_EN_INTERVIEW = {
    "research_background": "Why is this topic worth studying in the first place?",
    "research_question": "What is the central research question of the paper?",
    "core_conclusion": "What is the main conclusion of the paper?",
    "mechanism_analysis": "What mechanism does the author use to explain the result?",
    "policy_implication": "What policy relevance do you see in this paper?",
}
MODULE_QUESTIONS_CN_WRITTEN = {
    "research_background": "请概括本文的研究背景。",
    "research_question": "请说明本文的核心研究问题。",
    "core_conclusion": "请概括本文的核心结论。",
    "mechanism_analysis": "请分析本文的作用机制。",
    "policy_implication": "请说明本文的政策启示。",
}
MODULE_QUESTIONS_EN_WRITTEN = {
    "research_background": "Summarize the research background of the paper.",
    "research_question": "State the central research question of the paper.",
    "core_conclusion": "Summarize the core conclusion of the paper.",
    "mechanism_analysis": "Explain the mechanism emphasized in the paper.",
    "policy_implication": "Discuss the policy implication of the paper.",
}
MODULE_WHY_MATTERS = {
    "research_background": "研究背景决定你能否先把论文讲顺，避免一开口就进入细节。",
    "research_question": "研究问题是导师判断你是否真正抓住论文主线的第一步。",
    "core_conclusion": "核心结论是复试中最容易被追问和要求压缩表达的部分。",
    "mechanism_analysis": "机制分析能体现你有没有从结论上升到解释层面。",
    "policy_implication": "政策启示可以把论文理解和现实问题联系起来，提升回答完整性。",
}
WRITTEN_ANSWER_TYPES_CN = {
    "research_background": "简答题",
    "research_question": "简答题",
    "core_conclusion": "简答题",
    "mechanism_analysis": "论述题",
    "policy_implication": "论述题",
}
WRITTEN_ANSWER_TYPES_EN = {
    "research_background": "short-answer",
    "research_question": "short-answer",
    "core_conclusion": "short-answer",
    "mechanism_analysis": "analytical-response",
    "policy_implication": "analytical-response",
}

FIELD_KEYWORDS = {
    "数字经济学": [
        "digital",
        "digitization",
        "digitisation",
        "internet",
        "platform",
        "data",
        "algorithm",
        "online",
        "privacy",
        "network neutrality",
        "artificial intelligence",
        "人工智能",
        "数字",
        "平台",
        "数据",
        "算法",
        "互联网",
        "隐私",
        "数据要素",
    ],
    "宏观经济学": [
        "gdp",
        "inflation",
        "monetary",
        "fiscal",
        "business cycle",
        "growth",
        "失业",
        "通胀",
        "货币政策",
        "财政政策",
        "宏观",
        "经济增长",
    ],
    "微观经济学": [
        "consumer",
        "firm",
        "market",
        "pricing",
        "competition",
        "search cost",
        "price dispersion",
        "价格",
        "消费者",
        "企业",
        "市场",
        "竞争",
        "微观",
    ],
}
FIELD_ENGLISH = {
    "数字经济学": "digital economics",
    "宏观经济学": "macroeconomics",
    "微观经济学": "microeconomics",
    "相关经济学交叉方向": "interdisciplinary economics",
}

PAPER_TYPE_KEYWORDS = {
    "review": [
        "review",
        "survey",
        "literature",
        "journal of economic literature",
        "综述",
        "文献述评",
        "文献综述",
        "研究评述",
    ],
    "empirical": [
        "panel data",
        "evidence",
        "empirical",
        "difference-in-differences",
        "did",
        "regression",
        "sample",
        "based on",
        "基于",
        "实证",
        "面板数据",
        "样本",
        "回归",
        "异质性",
    ],
    "theoretical": [
        "model",
        "theory",
        "proposition",
        "assumption",
        "framework",
        "equilibrium",
        "模型",
        "理论",
        "命题",
        "假设",
        "均衡",
    ],
    "policy": [
        "policy",
        "regulation",
        "governance",
        "privacy",
        "copyright",
        "policy implications",
        "政策",
        "监管",
        "治理",
        "制度",
        "隐私",
        "版权",
    ],
}

DIGITAL_COST_LABELS = [
    ("search", "搜索成本"),
    ("replication", "复制成本"),
    ("transportation", "运输成本"),
    ("tracking", "追踪成本"),
    ("verification", "验证成本"),
]

KEYWORD_FALLBACKS = {
    "数字经济学": [
        "digital economics",
        "digital technology",
        "platform governance",
        "data",
        "privacy",
    ],
    "宏观经济学": [
        "economic growth",
        "inflation",
        "monetary policy",
        "fiscal policy",
        "business cycle",
    ],
    "微观经济学": [
        "market structure",
        "consumer behavior",
        "pricing",
        "competition",
        "search costs",
    ],
    "相关经济学交叉方向": [
        "economics",
        "policy",
        "institutions",
        "welfare",
        "technology",
    ],
}

POLICY_TERMS = {
    "network neutrality": "网络中立",
    "privacy": "隐私保护",
    "copyright": "版权治理",
    "discrimination": "反歧视",
    "reputation": "声誉机制",
    "platform": "平台治理",
    "net neutrality": "网络中立",
    "governance": "平台治理",
    "regulation": "监管规则",
    "data property": "数据产权",
}

TERM_TRANSLATIONS = {
    "数字经济": "digital economics",
    "数字技术": "digital technology",
    "平台治理": "platform governance",
    "数据要素流通": "data factor circulation",
    "区域创新": "regional innovation",
    "创新绩效": "innovation performance",
    "数字基础设施": "digital infrastructure",
    "隐私保护": "privacy protection",
    "知识溢出": "knowledge spillovers",
    "信息不对称": "information asymmetry",
    "资源配置": "resource allocation",
    "数据产权": "data property rights",
    "网络中立": "net neutrality",
    "版权治理": "copyright governance",
    "监管规则": "regulatory rules",
    "搜索成本": "search costs",
    "复制成本": "replication costs",
    "运输成本": "transportation costs",
    "追踪成本": "tracking costs",
    "验证成本": "verification costs",
    "宏观经济学": "macroeconomics",
    "微观经济学": "microeconomics",
    "数字经济学": "digital economics",
    "相关经济学交叉方向": "interdisciplinary economics",
}
TERM_EXPLANATIONS = {
    "数字经济": "强调数字技术如何改变资源配置、交易成本和市场组织方式。",
    "平台治理": "指平台规则、激励设计与责任分配的治理安排。",
    "数据要素流通": "指数据在不同主体之间的获取、流动、交易和使用过程。",
    "区域创新": "强调地区层面的创新产出、创新质量与创新效率变化。",
    "数字基础设施": "指支撑数字化生产、交易与治理的网络和算力基础。",
    "信息不对称": "指交易双方掌握的信息不对等，从而影响决策和资源配置。",
    "知识溢出": "指知识扩散到其他主体并产生外部收益的过程。",
    "资源配置": "指资本、劳动、数据等资源在不同部门和主体之间的配置效率。",
    "搜索成本": "指为获取信息、比较对象和寻找交易匹配所付出的成本。",
    "复制成本": "指复制信息产品或数字产品的边际成本。",
    "运输成本": "指商品、服务或信息在空间上转移的成本。",
    "追踪成本": "指识别、记录和监测交易行为及用户行为的成本。",
    "验证成本": "指确认交易对象、信息真实性和履约情况的成本。",
    "数字经济学": "研究数字技术和数据要素如何改变经济活动、市场结构与治理安排。",
}

EXPLICIT_TITLE_PATTERNS = [
    r"(?:^|\n)标题[:：]\s*(.+)",
    r"(?:^|\n)Title[:：]\s*(.+)",
]
EXPLICIT_AUTHOR_PATTERNS = [
    r"(?:^|\n)作者[:：]\s*(.+)",
    r"(?:^|\n)Authors?[:：]\s*(.+)",
]
EXPLICIT_YEAR_PATTERNS = [
    r"(?:^|\n)(?:年份|Year)[:：]\s*((?:19|20)\d{2})",
    r"\b((?:19|20)\d{2})\b",
]
EXPLICIT_KEYWORDS_PATTERNS = [
    r"(?:Keywords?|关键词)[:：]\s*(.+)",
]

ENGLISH_SECTION_HEADINGS = {
    "abstract": {"abstract"},
    "keywords": {"keywords", "jel", "jel classification"},
    "introduction": {"introduction"},
    "conclusion": {"conclusion", "conclusions", "concluding remarks", "concluding comments"},
    "references": {"references", "bibliography"},
}
CHINESE_SECTION_HEADINGS = {
    "abstract": {"摘要"},
    "keywords": {"关键词"},
    "introduction": {"引言", "导论", "绪论", "前言"},
    "conclusion": {"结论", "研究结论", "结语", "总结"},
    "references": {"参考文献"},
}

PRINTED_OUTPUTS = [
    "full.json",
    "general_interview.json",
    "english_interview.json",
    "professional_written_exam.json",
    "english_written_exam.json",
    "retest_pack.xlsx",
    "retest_pack_memorize.xlsx",
    "retest_pack_print.xlsx",
]


@dataclass
class UserFacingError(Exception):
    """Message that should be shown directly to the user."""

    message: str
    code: int

    def __str__(self) -> str:
        return self.message


@dataclass
class SourceData:
    input_type: str
    raw_text: str
    source_path: str
    page_count: int | None = None
    non_empty_page_ratio: float | None = None


@dataclass
class PaperMetadata:
    title: str
    authors: list[str]
    year: str
    language: str
    language_confidence: str
    language_reason: str
    field: str
    keywords: list[str]
    paper_type: str


@dataclass
class PaperSections:
    abstract: str
    abstract_sentences: list[str]
    conclusion: str
    conclusion_sentences: list[str]
    introduction: str
    intro_sentences: list[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate standardized retest outputs from a paper.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--input-file", type=Path, help="Path to a PDF, TXT, or MD file.")
    group.add_argument("--input-text", help="Inline text content.")
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Output root directory. Defaults to ./output/economics-retest-paper-splitter",
    )
    parser.add_argument("--slug", help="Override the automatically generated paper slug.")
    parser.add_argument(
        "--language",
        choices=sorted(ALLOWED_LANGUAGES),
        help="Force the paper language instead of auto-detecting it.",
    )
    parser.add_argument(
        "--stdout-json",
        action="store_true",
        help="Print full.json content to stdout instead of writing files.",
    )
    return parser.parse_args()


def clean_text(text: str) -> str:
    text = text.replace("\u00a0", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def load_source(input_file: Path | None, input_text: str | None) -> SourceData:
    if input_text is not None:
        return SourceData(input_type="text", raw_text=clean_text(input_text), source_path="<inline-text>")
    assert input_file is not None
    suffix = input_file.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf_source(input_file)
    text = input_file.read_text(encoding="utf-8")
    return SourceData(input_type=suffix.lstrip(".") or "text", raw_text=clean_text(text), source_path=str(input_file.resolve()))


def extract_pdf_source(path: Path) -> SourceData:
    reader = PdfReader(str(path))
    page_texts = [(page.extract_text() or "").strip() for page in reader.pages]
    joined = clean_text("\n".join(text for text in page_texts if text))
    non_empty_pages = sum(1 for text in page_texts if len(text) >= 80)
    ratio = non_empty_pages / max(1, len(page_texts))
    if len(joined) < MIN_PDF_TEXT_LENGTH or ratio < MIN_NON_EMPTY_PAGE_RATIO:
        raise UserFacingError(STANDARD_TEXT_PROMPT, EXIT_PDF_EXTRACTION_FAILED)
    return SourceData(
        input_type="pdf",
        raw_text=joined,
        source_path=str(path.resolve()),
        page_count=len(page_texts),
        non_empty_page_ratio=ratio,
    )


def normalized_heading(line: str) -> str:
    line = clean_text(line).strip(":：")
    line = re.sub(r"^[\d一二三四五六七八九十IVXivx.()（）\s-]+", "", line)
    return line.lower()


def is_heading_line(line: str, language: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    norm = normalized_heading(stripped)
    english_heads = set().union(*ENGLISH_SECTION_HEADINGS.values())
    chinese_heads = set().union(*CHINESE_SECTION_HEADINGS.values())
    if norm in english_heads or stripped in chinese_heads:
        return True
    if language == "中文文献":
        return bool(re.fullmatch(r"[一二三四五六七八九十]+[、.．].{1,25}", stripped))
    return bool(re.fullmatch(r"\d+(?:\.\d+)*\s+[A-Z].{0,80}", stripped))


def split_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines()]


def detect_language(text: str, title_hint: str = "", abstract_hint: str = "", forced: str | None = None) -> tuple[str, str, str]:
    if forced:
        return forced, "forced", "使用命令行参数强制指定文献语言。"

    def score(chunk: str) -> tuple[int, int]:
        chinese = len(re.findall(r"[\u4e00-\u9fff]", chunk))
        latin = len(re.findall(r"[A-Za-z]{2,}", chunk))
        return chinese, latin

    priority_text = clean_text("\n".join(part for part in [title_hint, abstract_hint] if part))
    body_sample = clean_text(text[:12000])
    priority_chinese, priority_latin = score(priority_text)
    body_chinese, body_latin = score(body_sample)

    if priority_text:
        if priority_chinese >= 8 and priority_latin >= 12:
            return "中英混合文献", "medium", "标题和摘要同时包含较多中文与英文内容。"
        if priority_chinese >= 12 and priority_latin < 35:
            return "中文文献", "high", "标题与摘要主体以中文为主。"
        if priority_latin >= 20 and priority_chinese < 8:
            return "英文文献", "high", "标题与摘要主体以英文连续文本为主。"

    if body_chinese >= 25 and body_latin >= 25:
        return "中英混合文献", "medium", "正文抽样同时包含较多中文和英文内容。"
    if body_chinese >= 70 and body_latin < 120:
        return "中文文献", "medium", "正文抽样以中文内容为主。"
    if body_latin >= 70 and body_chinese < 30:
        return "英文文献", "medium", "正文抽样以英文内容为主。"
    raise UserFacingError(STANDARD_LANGUAGE_PROMPT, EXIT_LANGUAGE_UNKNOWN)


def find_labeled_value(text: str, patterns: list[str], flags: int = re.IGNORECASE) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return clean_text(match.group(1))
    return ""


def extract_title(text: str, input_path: Path | None) -> tuple[str, bool]:
    explicit = find_labeled_value(text, EXPLICIT_TITLE_PATTERNS)
    if explicit:
        return explicit, False

    lines = [line for line in split_lines(text) if line]
    skip_patterns = [
        r"working paper",
        r"series",
        r"http",
        r"journal",
        r"nber",
        r"massachusetts avenue",
        r"abstract",
        r"jel",
        r"keywords?",
        r"references",
        r"摘要",
        r"关键词",
        r"引言",
        r"结论",
        r"参考文献",
    ]
    for line in lines[:50]:
        lowered = line.lower()
        if len(line) < 4 or len(line) > 180:
            continue
        if any(re.search(pattern, lowered) for pattern in skip_patterns):
            continue
        if re.fullmatch(r"[\d\W_]+", line):
            continue
        if re.search(r"[A-Za-z\u4e00-\u9fff]", line):
            return normalize_title_case(line), False

    if input_path is not None:
        return normalize_title_case(input_path.stem.replace("-", " ")), True
    return "Untitled Paper", True


def normalize_title_case(title: str) -> str:
    words = title.split()
    if not words:
        return title
    if all(word.isupper() for word in words if re.search(r"[A-Z]", word)):
        return " ".join(word.capitalize() for word in words)
    return title


def extract_authors(text: str, title: str) -> list[str]:
    labeled = find_labeled_value(text, EXPLICIT_AUTHOR_PATTERNS)
    if labeled:
        parts = [part.strip() for part in re.split(r"[,;；，、]| and ", labeled) if part.strip()]
        return parts[:6]

    lines = [line for line in split_lines(text) if line]
    title_candidates = {title, title.replace(" ", "")}
    name_pattern = re.compile(r"^[A-Z][A-Za-z.'-]+(?:\s+[A-Z][A-Za-z.'-]+){1,3}$")
    for index, line in enumerate(lines[:60]):
        if line in title_candidates:
            authors: list[str] = []
            for candidate_line in lines[index + 1 : index + 8]:
                candidate = candidate_line.replace("∗", "").replace("†", "").strip()
                if name_pattern.fullmatch(candidate):
                    authors.append(candidate)
            if authors:
                return authors[:6]
    return []


def extract_year(text: str) -> str:
    labeled = find_labeled_value(text, EXPLICIT_YEAR_PATTERNS)
    if labeled:
        return labeled[:4]
    return ""


def extract_section_block(text: str, start_heads: set[str], stop_heads: set[str], language: str) -> str:
    lines = split_lines(text)
    start_index: int | None = None
    for index, line in enumerate(lines):
        if not line:
            continue
        norm = normalized_heading(line)
        if norm in start_heads or line in start_heads:
            start_index = index + 1
            break
    if start_index is None:
        return ""

    collected: list[str] = []
    for line in lines[start_index:]:
        if not line:
            if collected and collected[-1]:
                collected.append("")
            continue
        norm = normalized_heading(line)
        if norm in stop_heads or line in stop_heads:
            break
        if collected and is_heading_line(line, language) and len(" ".join(collected)) > 120:
            break
        collected.append(line)
    return clean_text("\n".join(collected))


def extract_abstract(text: str, language: str) -> str:
    if language == "中文文献":
        abstract = extract_section_block(
            text,
            CHINESE_SECTION_HEADINGS["abstract"],
            CHINESE_SECTION_HEADINGS["keywords"] | CHINESE_SECTION_HEADINGS["introduction"],
            language,
        )
        if abstract:
            return abstract
        fallback = re.search(r"(?:摘要)[:：]?\s*(.+?)(?:\n\s*关键词|\n\s*引言|\Z)", text, re.DOTALL)
        if fallback:
            return clean_text(fallback.group(1))
    else:
        abstract = extract_section_block(
            text,
            ENGLISH_SECTION_HEADINGS["abstract"],
            ENGLISH_SECTION_HEADINGS["keywords"] | ENGLISH_SECTION_HEADINGS["introduction"],
            "英文文献",
        )
        if abstract:
            return abstract
        fallback = re.search(
            r"(?:ABSTRACT|Abstract)[:：]?\s*(.+?)(?:\n\s*Keywords|\n\s*JEL|\n\s*(?:1\s+Introduction|Introduction)|\Z)",
            text,
            re.IGNORECASE | re.DOTALL,
        )
        if fallback:
            return clean_text(fallback.group(1))

    paragraphs = [paragraph.strip() for paragraph in text.split("\n\n") if paragraph.strip()]
    if paragraphs:
        return clean_text(paragraphs[0])[:2400]
    return ""


def extract_conclusion(text: str, language: str) -> str:
    if language == "中文文献":
        conclusion = extract_section_block(
            text,
            CHINESE_SECTION_HEADINGS["conclusion"],
            CHINESE_SECTION_HEADINGS["references"],
            language,
        )
        if conclusion:
            return conclusion[:2500]
        fallback = re.search(r"(?:结论|结语)[:：]?\s*(.+?)(?:\n\s*参考文献|\Z)", text, re.DOTALL)
        if fallback:
            return clean_text(fallback.group(1))[:2500]
    else:
        conclusion = extract_section_block(
            text,
            ENGLISH_SECTION_HEADINGS["conclusion"],
            ENGLISH_SECTION_HEADINGS["references"],
            "英文文献",
        )
        if conclusion:
            return conclusion[:2500]
        fallback = re.search(
            r"(?:Conclusion|Conclusions|Concluding Remarks)[:：]?\s*(.+?)(?:\n\s*References|\Z)",
            text,
            re.IGNORECASE | re.DOTALL,
        )
        if fallback:
            return clean_text(fallback.group(1))[:2500]
    return ""


def extract_introduction(text: str, language: str) -> str:
    if language == "中文文献":
        intro = extract_section_block(
            text,
            CHINESE_SECTION_HEADINGS["introduction"],
            CHINESE_SECTION_HEADINGS["conclusion"] | CHINESE_SECTION_HEADINGS["references"],
            language,
        )
        return intro[:2200]
    intro = extract_section_block(
        text,
        ENGLISH_SECTION_HEADINGS["introduction"],
        ENGLISH_SECTION_HEADINGS["conclusion"] | ENGLISH_SECTION_HEADINGS["references"],
        "英文文献",
    )
    return intro[:2200]


def split_sentences(text: str, language: str) -> list[str]:
    if not text:
        return []
    if language == "中文文献":
        parts = re.split(r"(?<=[。！？；])", text)
    else:
        parts = re.split(r"(?<=[.!?])\s+", text)
    return [clean_text(part) for part in parts if clean_text(part)]


def determine_field(title: str, abstract: str, conclusion: str) -> str:
    haystack = f"{title}\n{abstract}\n{conclusion}".lower()
    scores = {}
    for field, keywords in FIELD_KEYWORDS.items():
        scores[field] = sum(1 for keyword in keywords if keyword in haystack)
    best_field = max(scores, key=scores.get)
    return best_field if scores[best_field] > 0 else "相关经济学交叉方向"


def extract_keywords(title: str, abstract: str, text: str, field: str, language: str) -> list[str]:
    explicit = find_labeled_value(text, EXPLICIT_KEYWORDS_PATTERNS)
    if explicit:
        parts = [part.strip() for part in re.split(r"[,;；，、]", explicit) if part.strip()]
        if parts:
            return parts[:8]

    haystack = f"{title}\n{abstract}".lower()
    if language == "中文文献":
        title_terms = [
            term.strip()
            for term in re.split(r"[、，；：:：\-—\s与和]+", re.sub(r"[^\u4e00-\u9fffA-Za-z0-9、，；：:：\-—\s与和]", " ", title))
            if len(term.strip()) >= 2
        ]
        phrase_candidates = [
            "数据要素流通",
            "平台治理",
            "区域创新",
            "创新绩效",
            "数字基础设施",
            "隐私保护",
            "知识溢出",
            "信息不对称",
            "资源配置",
            "政府治理能力",
            "市场化水平",
        ]
        found_cn: list[str] = []
        for candidate in title_terms + phrase_candidates:
            if candidate in f"{title}\n{abstract}" and candidate not in found_cn:
                found_cn.append(candidate)
        if found_cn:
            return found_cn[:8]

    found: list[str] = []
    for keyword in KEYWORD_FALLBACKS[field]:
        if keyword.lower() in haystack and keyword not in found:
            found.append(keyword)
    if found:
        return found[:8]
    return KEYWORD_FALLBACKS[field][:5]


def slugify(title: str) -> tuple[str, bool]:
    title = title.lower()
    slug = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "-", title)
    slug = re.sub(r"-{2,}", "-", slug).strip("-")
    if slug:
        return slug, False
    return "untitled-paper", True


def detect_cost_terms(text: str) -> list[str]:
    lowered = text.lower()
    labels = [label for token, label in DIGITAL_COST_LABELS if token in lowered]
    deduped: list[str] = []
    for label in labels:
        if label not in deduped:
            deduped.append(label)
    return deduped


def detect_policy_topics(text: str) -> list[str]:
    lowered = text.lower()
    topics = [label for token, label in POLICY_TERMS.items() if token in lowered]
    deduped: list[str] = []
    for topic in topics:
        if topic not in deduped:
            deduped.append(topic)
    return deduped[:6]


def detect_paper_type(title: str, abstract: str, conclusion: str, text: str) -> str:
    haystack = f"{title}\n{abstract}\n{conclusion}\n{text[:5000]}".lower()
    scores = {paper_type: 0 for paper_type in PAPER_TYPE_KEYWORDS}
    for paper_type, keywords in PAPER_TYPE_KEYWORDS.items():
        scores[paper_type] = sum(1 for keyword in keywords if keyword in haystack)

    if scores["review"] > 0:
        return "review"
    if scores["empirical"] >= 2:
        return "empirical"
    if scores["theoretical"] >= 2:
        return "theoretical"
    if scores["policy"] >= 2:
        return "policy"

    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "empirical"


def chinese_anchor_from_english(english_sentence: str, fallback: str) -> str:
    if re.search(r"[\u4e00-\u9fff]", english_sentence):
        return english_sentence
    lowered = english_sentence.lower()
    if "search" in lowered and "cost" in lowered:
        return "文章强调搜索成本下降会改变信息比较、匹配效率和价格形成。"
    if "cost" in lowered:
        return "文章强调数字化会通过重塑关键经济成本结构来改变市场行为和经济结果。"
    if "privacy" in lowered:
        return "文章指出隐私与数据使用边界会直接影响数字市场的运行规则。"
    if "productivity" in lowered:
        return "文章强调数字技术对生产率的影响存在明显异质性，并取决于组织能力。"
    if "policy" in lowered or "regulation" in lowered:
        return "文章提示数字经济研究的政策含义主要体现在监管、治理与制度边界上。"
    if "review" in lowered or "literature" in lowered:
        return "文章通过综述方式梳理相关文献，并尝试提炼统一分析框架。"
    return fallback


def field_to_english(field: str) -> str:
    return FIELD_ENGLISH.get(field, "economics")


def module_label(module: str) -> str:
    return MODULE_LABELS[module]


def pick_sentence(sentences: list[str], patterns: list[str]) -> str:
    for sentence in sentences:
        lowered = sentence.lower()
        if any(pattern in lowered for pattern in patterns):
            return clean_text(sentence)
    return ""


def inferred_text(text: str) -> str:
    return f"原文未充分展开，根据摘要/正文可推断为：{text}"


def infer_background(metadata: PaperMetadata, sections: PaperSections) -> str:
    if metadata.language == "中文文献":
        direct = pick_sentence(sections.intro_sentences + sections.abstract_sentences, ["背景", "随着", "近年来", "数字", "平台", "数据"])
        if direct:
            return direct
        if sections.abstract_sentences:
            return sections.abstract_sentences[0]
    else:
        if sections.abstract_sentences:
            return chinese_anchor_from_english(
                sections.abstract_sentences[0],
                "文章关注数字化条件变化如何重塑经济活动与制度环境。",
            )

    fallback = {
        "review": "文章试图解释数字技术为什么会系统性改变经济活动，而不是把数字经济当成与传统经济学完全割裂的新问题。",
        "empirical": "文章关注某一数字化因素是否真实影响了经济结果，以及这种影响为何会在不同条件下发生差异。",
        "theoretical": "文章试图通过模型分析澄清数字化情境下的关键机制和边界条件。",
        "policy": "文章关注数字化带来的效率提升与制度治理之间如何形成新的政策张力。",
    }[metadata.paper_type]
    return inferred_text(fallback)


def infer_question(metadata: PaperMetadata, sections: PaperSections) -> str:
    if metadata.language == "中文文献":
        for sentence in sections.abstract_sentences + sections.intro_sentences:
            if any(token in sentence for token in ["考察", "研究", "讨论", "分析", "检验", "评估"]):
                return sentence
    else:
        direct = pick_sentence(sections.abstract_sentences + sections.intro_sentences, ["examines", "studies", "asks whether", "investigates"])
        if direct:
            return chinese_anchor_from_english(
                direct,
                f"文章围绕《{metadata.title}》所对应的问题，讨论数字化变化如何影响经济活动。",
            )

    fallback = {
        "review": f"文章的核心问题是：如何用统一的经济学框架解释《{metadata.title}》所涉及的数字化现象。",
        "empirical": "文章的核心问题是：特定数字化因素是否显著影响目标经济结果，以及这种影响是否具有异质性。",
        "theoretical": "文章的核心问题是：在给定假设下，数字化变化会通过什么机制改变行为选择与均衡结果。",
        "policy": "文章的核心问题是：制度规则和治理安排如何影响数字经济中的效率、创新与风险分配。",
    }[metadata.paper_type]
    return inferred_text(fallback)


def infer_conclusion(metadata: PaperMetadata, sections: PaperSections) -> str:
    if metadata.language == "中文文献":
        direct = pick_sentence(sections.conclusion_sentences + sections.abstract_sentences, ["研究发现", "结果表明", "说明", "发现", "表明"])
        if direct:
            return direct
    else:
        direct = pick_sentence(sections.conclusion_sentences + sections.abstract_sentences, ["find", "show", "conclude", "core theme"])
        if direct:
            return chinese_anchor_from_english(
                direct,
                "文章认为数字化会显著改变相关经济结果，但作用强度取决于制度环境和行为约束。",
            )

    fallback = {
        "review": "文章认为，数字经济的许多现象仍可用标准经济学解释，但分析重点必须转向成本结构变化。",
        "empirical": "文章认为，数字化因素会显著影响目标经济结果，但影响强度受异质性和制度条件约束。",
        "theoretical": "文章认为，理论结论依赖关键假设，因此结论的外部适用性需要结合现实条件判断。",
        "policy": "文章认为，数字经济政策不能只追求效率，还需要兼顾制度约束、治理成本与分配后果。",
    }[metadata.paper_type]
    return inferred_text(fallback)


def infer_mechanism(metadata: PaperMetadata, sections: PaperSections, cost_terms: list[str]) -> str:
    if metadata.language == "中文文献":
        direct = pick_sentence(sections.abstract_sentences + sections.conclusion_sentences, ["机制", "通过", "路径", "信息不对称", "知识溢出", "资源配置"])
        if direct:
            return direct
    else:
        direct = pick_sentence(sections.abstract_sentences + sections.conclusion_sentences, ["through", "mechanism", "cost", "information", "spillover"])
        if direct:
            return chinese_anchor_from_english(
                direct,
                "文章强调数字化主要通过改变信息、匹配与治理成本来影响经济结果。",
            )

    if cost_terms:
        return inferred_text(f"文章可被理解为：数字化通过改变{'、'.join(cost_terms)}等关键成本，进而改变资源配置、激励结构与市场结果。")

    fallback = {
        "review": "文章可被理解为：数字化通过改变信息获取、复制传播和交易验证等成本，重新塑造市场与组织行为。",
        "empirical": "文章可被理解为：数字化因素通过缓解信息不对称、改善资源配置或强化知识扩散来影响结果变量。",
        "theoretical": "文章可被理解为：关键假设变化会通过激励结构调整传导到行为选择和均衡结果。",
        "policy": "文章可被理解为：制度规则通过改变约束条件和治理激励来影响市场行为与政策绩效。",
    }[metadata.paper_type]
    return inferred_text(fallback)


def infer_policy(metadata: PaperMetadata, sections: PaperSections, policy_topics: list[str]) -> str:
    if metadata.language == "中文文献":
        direct = pick_sentence(sections.abstract_sentences + sections.conclusion_sentences, ["政策", "监管", "制度", "启示", "隐私", "版权"])
        if direct:
            return direct
        direct = pick_sentence(sections.abstract_sentences + sections.conclusion_sentences, ["治理"])
        if direct:
            return direct
    else:
        direct = pick_sentence(sections.abstract_sentences + sections.conclusion_sentences, ["policy", "regulation", "privacy", "copyright", "net neutrality"])
        if direct:
            return chinese_anchor_from_english(
                direct,
                "文章的政策含义主要落在平台治理、隐私保护和制度规则设计上。",
            )
        direct = pick_sentence(sections.abstract_sentences + sections.conclusion_sentences, ["governance"])
        if direct:
            return chinese_anchor_from_english(
                direct,
                "文章的政策含义主要落在平台治理、隐私保护和制度规则设计上。",
            )

    if policy_topics:
        return inferred_text(f"文章提示政策设计需要重点关注{'、'.join(policy_topics)}等制度问题，并兼顾效率与治理目标。")

    fallback = {
        "review": "文章提示政策讨论不能停留在技术扩散层面，而应同步处理治理、规则与市场边界。",
        "empirical": "文章提示政策效果依赖制度环境，因此推进数字化时要重视治理能力、监管规则和配套制度。",
        "theoretical": "文章提示政策判断不能脱离假设条件，需要结合现实制度和实施约束解释理论结论。",
        "policy": "文章提示政策设计必须平衡创新激励、治理成本和风险约束，不能单看效率提升。",
    }[metadata.paper_type]
    return inferred_text(fallback)


def build_mandatory_blocks(metadata: PaperMetadata, sections: PaperSections, cost_terms: list[str], policy_topics: list[str]) -> dict[str, str]:
    return {
        "research_background": infer_background(metadata, sections),
        "research_question": infer_question(metadata, sections),
        "core_conclusion": infer_conclusion(metadata, sections),
        "mechanism_analysis": infer_mechanism(metadata, sections, cost_terms),
        "policy_implication": infer_policy(metadata, sections, policy_topics),
    }


def build_one_sentence_summary(metadata: PaperMetadata, mandatory_blocks: dict[str, str]) -> str:
    question = mandatory_blocks["research_question"].replace("原文未充分展开，根据摘要/正文可推断为：", "")
    conclusion = mandatory_blocks["core_conclusion"].replace("原文未充分展开，根据摘要/正文可推断为：", "")
    if metadata.paper_type == "review":
        return f"本文围绕{question}展开，并认为数字经济研究应在成本结构变化这一统一框架下理解。"
    return f"本文围绕{question}展开，核心结论是{conclusion}"


def lookup_english_term(chinese_term: str) -> str:
    if chinese_term in TERM_TRANSLATIONS:
        return TERM_TRANSLATIONS[chinese_term]
    if re.search(r"[A-Za-z]", chinese_term):
        return chinese_term
    return chinese_term


def lookup_chinese_term(english_term: str) -> str:
    lowered = english_term.lower()
    for chinese, english in TERM_TRANSLATIONS.items():
        if english.lower() == lowered:
            return chinese
    return english_term


def build_terms(metadata: PaperMetadata, cost_terms: list[str], policy_topics: list[str]) -> list[dict[str, str]]:
    seeds = list(metadata.keywords) + cost_terms + policy_topics + [metadata.field]
    seen: set[str] = set()
    terms: list[dict[str, str]] = []
    for term in seeds:
        if not term:
            continue
        if re.search(r"[A-Za-z]", term) and not re.search(r"[\u4e00-\u9fff]", term):
            english = term
            chinese = lookup_chinese_term(term)
        else:
            chinese = term
            english = lookup_english_term(term)
        key = f"{chinese}|{english}"
        if key in seen:
            continue
        seen.add(key)
        explanation = TERM_EXPLANATIONS.get(chinese) or TERM_EXPLANATIONS.get(english) or f"与{module_label('mechanism_analysis')}或论文主题直接相关的术语。"
        terms.append({"chinese": chinese, "english": english, "explanation": explanation})
        if len(terms) >= 8:
            break
    return terms


def build_general_answer(module: str, content: str) -> str:
    prefix = {
        "research_background": "如果口头概括这篇论文的研究背景，我会先说：",
        "research_question": "这篇论文真正要回答的问题可以概括为：",
        "core_conclusion": "如果导师让我直接说结论，我会回答：",
        "mechanism_analysis": "就机制而言，我会把这篇论文理解为：",
        "policy_implication": "从政策层面看，我会强调：",
    }[module]
    return f"{prefix}{content}"


def build_general_interview(mandatory_blocks: dict[str, str]) -> list[dict[str, str]]:
    return [
        {
            "module": module,
            "question": MODULE_QUESTIONS_CN[module],
            "answer": build_general_answer(module, mandatory_blocks[module]),
            "why_this_matters": MODULE_WHY_MATTERS[module],
        }
        for module in MANDATORY_MODULES
    ]


def english_term_note(terms: list[dict[str, str]], limit: int = 3) -> str:
    selected = [f"{item['english']} ({item['chinese']})" for item in terms[:limit]]
    return "; ".join(selected)


def build_english_interview_answer(module: str, metadata: PaperMetadata, mandatory_blocks: dict[str, str], terms: list[dict[str, str]]) -> str:
    field_en = field_to_english(metadata.field)
    core_terms = english_term_note(terms, limit=3)
    if module == "research_background":
        if metadata.paper_type == "review":
            return "The paper starts from the observation that digital technologies are reshaping economic activity, so the author tries to organize scattered studies into a coherent economic framework."
        if metadata.paper_type == "empirical":
            return f"The paper is motivated by a practical question in {field_en}: whether a specific digital factor has a measurable effect on the target outcome and why the effect differs across settings."
        if metadata.paper_type == "theoretical":
            return f"The background is mainly theoretical. The author wants to clarify how digitalization changes incentives and equilibrium conditions in {field_en}."
        return f"The paper is motivated by the policy tension between digital efficiency gains and institutional governance in {field_en}."
    if module == "research_question":
        return f"At its core, the paper asks how the focal digital change affects economic behavior, market outcomes, and institutional arrangements in {field_en}."
    if module == "core_conclusion":
        if metadata.paper_type == "review":
            return "The main conclusion is that many digital-economy phenomena can still be explained by standard economics once we focus on how digitization changes cost structures."
        if metadata.paper_type == "empirical":
            return "The paper concludes that the focal digital factor matters for the outcome variable, but the strength of the effect depends on heterogeneity and institutional conditions."
        if metadata.paper_type == "theoretical":
            return "The paper concludes that the predicted result depends on key assumptions, so the mechanism is clearer than the universal applicability of the conclusion."
        return "The paper concludes that digital policy should not be judged by efficiency alone, because governance constraints and distributional effects matter as well."
    if module == "mechanism_analysis":
        return f"The mechanism is explained through changes in costs, incentives, and information conditions. In practical terms, the paper links the result to terms such as {core_terms}."
    return "The policy implication is that digital development should be accompanied by stronger institutional design, especially when governance quality, regulation, and implementation capacity shape the final outcome."


def build_english_interview(mandatory_blocks: dict[str, str], metadata: PaperMetadata, terms: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        {
            "module": module,
            "question_en": MODULE_QUESTIONS_EN_INTERVIEW[module],
            "answer_en": build_english_interview_answer(module, metadata, mandatory_blocks, terms),
            "terminology_notes": english_term_note(terms, limit=3),
        }
        for module in MANDATORY_MODULES
    ]


def build_written_answer(module: str, mandatory_blocks: dict[str, str]) -> str:
    content = mandatory_blocks[module]
    if content.startswith("原文未充分展开"):
        return content
    prefix = {
        "research_background": "从研究背景看，",
        "research_question": "从研究问题看，",
        "core_conclusion": "从核心结论看，",
        "mechanism_analysis": "从作用机制看，",
        "policy_implication": "从政策启示看，",
    }[module]
    return f"{prefix}{content}"


def build_professional_written_exam(mandatory_blocks: dict[str, str]) -> list[dict[str, str]]:
    return [
        {
            "module": module,
            "question": MODULE_QUESTIONS_CN_WRITTEN[module],
            "answer": build_written_answer(module, mandatory_blocks),
            "answer_type": WRITTEN_ANSWER_TYPES_CN[module],
        }
        for module in MANDATORY_MODULES
    ]


def build_english_written_answer(module: str, metadata: PaperMetadata, terms: list[dict[str, str]]) -> str:
    field_en = field_to_english(metadata.field)
    key_term = terms[0]["english"] if terms else "digitalization"
    if module == "research_background":
        return f"The research background lies in the growing importance of {key_term} in {field_en} and in the need to explain its economic consequences in a structured way."
    if module == "research_question":
        return f"The paper asks how the focal digital change influences economic outcomes, behavioral incentives, and institutional arrangements in {field_en}."
    if module == "core_conclusion":
        if metadata.paper_type == "review":
            return "The paper argues that digital-economy phenomena can be interpreted within standard economic reasoning once the analysis is centered on changes in cost structures."
        if metadata.paper_type == "empirical":
            return "The paper finds that the focal digital variable has a meaningful effect on the target outcome, although the effect varies across contexts and institutions."
        if metadata.paper_type == "theoretical":
            return "The paper shows that the theoretical result depends on the model assumptions and should therefore be interpreted together with its boundary conditions."
        return "The paper argues that policy assessment in the digital economy must balance efficiency, governance, and implementation constraints."
    if module == "mechanism_analysis":
        return "The mechanism can be summarized as a chain from changing costs and information conditions to shifting incentives, resource allocation, and finally economic outcomes."
    return "The policy implication is that digital transformation needs complementary governance, regulation, and institutional capacity rather than a purely technology-driven approach."


def build_english_written_exam(metadata: PaperMetadata, terms: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        {
            "module": module,
            "question_en": MODULE_QUESTIONS_EN_WRITTEN[module],
            "answer_en": build_english_written_answer(module, metadata, terms),
            "answer_type": WRITTEN_ANSWER_TYPES_EN[module],
        }
        for module in MANDATORY_MODULES
    ]


def build_review_outline() -> dict[str, list[str]]:
    return {
        "general_interview_outline": [MODULE_LABELS[module] for module in MANDATORY_MODULES],
        "english_interview_outline": ["Research Background", "Research Question", "Core Conclusion", "Mechanism Analysis", "Policy Implication"],
        "professional_written_exam_outline": [MODULE_LABELS[module] for module in MANDATORY_MODULES],
        "english_written_exam_outline": ["Research Background", "Research Question", "Core Conclusion", "Mechanism Analysis", "Policy Implication"],
    }


def build_low_priority(metadata: PaperMetadata) -> list[dict[str, str]]:
    if metadata.paper_type == "review":
        return [
            {"label": "具体文献条目清单", "reason": "对复试主线帮助有限，优先级低于统一框架和政策含义。"},
            {"label": "机构与版本信息", "reason": "不影响对论文核心思想的表达。"},
        ]
    if metadata.paper_type == "empirical":
        return [
            {"label": "过细的数据清洗过程", "reason": "复试中迁移性弱，通常不应占据主干结构。"},
            {"label": "稳健性检验的附录细节", "reason": "除非导师专门追问，否则优先级低于机制与主结论。"},
        ]
    if metadata.paper_type == "theoretical":
        return [
            {"label": "附录级技术推导", "reason": "对第一轮口答和卷面作答帮助较弱。"},
            {"label": "复杂证明细节", "reason": "可作为深入阅读内容，但不宜替代固定模块。"},
        ]
    return [
        {"label": "表格级政策背景细节", "reason": "对标准化复试表达的帮助弱于机制与启示。"},
        {"label": "边缘案例说明", "reason": "适合作为补充阅读，不宜占据主框架。"},
    ]


def build_extra(metadata: PaperMetadata) -> dict[str, list[str]]:
    innovation = {
        "review": ["文章通过统一框架整合分散文献，而不是简单罗列已有观点。"],
        "empirical": ["文章把数字化变量的总体效应、机制检验和异质性分析放进同一实证框架中。"],
        "theoretical": ["文章通过模型设定澄清数字化条件下的机制和边界。"],
        "policy": ["文章把制度工具、实施约束和政策效果统一到一个分析框架中。"],
    }[metadata.paper_type]
    limitations = {
        "review": ["作为综述，文章更擅长框架整合，弱于对单个机制进行严格识别。"],
        "empirical": ["文章的外部有效性和识别强度仍然取决于样本范围与制度背景。"],
        "theoretical": ["理论结论对假设条件依赖较强，现实适用性需要进一步验证。"],
        "policy": ["政策判断可能受到制度环境差异和实施能力差异的影响。"],
    }[metadata.paper_type]
    extension = {
        "review": ["后续可继续比较不同数字场景下同一机制为何会呈现不同结果。"],
        "empirical": ["后续可继续识别机制的边界条件，并考察不同地区或组织条件下的差异。"],
        "theoretical": ["后续可进一步放松假设，并与经验研究结合检验理论边界。"],
        "policy": ["后续可继续比较不同政策工具组合对效率与治理目标的权衡。"],
    }[metadata.paper_type]
    return {
        "innovation_points": innovation,
        "limitations": limitations,
        "extension_research": extension,
        "source_notes": ["固定 5 模块为强制主框架，可选内容仅作为补充，不替代主干结构。"],
    }


def build_metadata(source: SourceData, forced_language: str | None = None) -> tuple[PaperMetadata, PaperSections, bool]:
    cleaned = source.raw_text
    if len(cleaned) < MIN_TEXT_LENGTH:
        raise UserFacingError(STANDARD_TEXT_PROMPT, EXIT_INFO_INSUFFICIENT)

    title, used_title_fallback = extract_title(cleaned, Path(source.source_path) if source.source_path != "<inline-text>" else None)
    authors = extract_authors(cleaned, title)
    year = extract_year(cleaned)

    provisional_language, _, _ = detect_language(cleaned, title_hint=title, abstract_hint="", forced=forced_language)
    abstract = extract_abstract(cleaned, provisional_language)
    language, confidence, reason = detect_language(cleaned, title_hint=title, abstract_hint=abstract, forced=forced_language)

    min_abstract_length = 60 if language == "中文文献" else 120
    if len(clean_text(abstract)) < min_abstract_length:
        raise UserFacingError(STANDARD_TEXT_PROMPT, EXIT_INFO_INSUFFICIENT)

    conclusion = extract_conclusion(cleaned, language)
    introduction = extract_introduction(cleaned, language)
    sections = PaperSections(
        abstract=abstract,
        abstract_sentences=split_sentences(abstract, language),
        conclusion=conclusion,
        conclusion_sentences=split_sentences(conclusion, language),
        introduction=introduction,
        intro_sentences=split_sentences(introduction, language),
    )
    field = determine_field(title, abstract, conclusion)
    keywords = extract_keywords(title, abstract, cleaned, field, language)
    paper_type = detect_paper_type(title, abstract, conclusion, cleaned)

    metadata = PaperMetadata(
        title=title,
        authors=authors,
        year=year,
        language=language,
        language_confidence=confidence,
        language_reason=reason,
        field=field,
        keywords=keywords,
        paper_type=paper_type,
    )
    return metadata, sections, used_title_fallback


def build_full_output(source: SourceData, forced_language: str | None = None) -> tuple[dict[str, Any], str, bool, PaperMetadata, PaperSections]:
    metadata, sections, used_title_fallback = build_metadata(source, forced_language=forced_language)
    combined_text = "\n".join([source.raw_text, sections.abstract, sections.conclusion])
    cost_terms = detect_cost_terms(combined_text)
    policy_topics = detect_policy_topics(combined_text)
    mandatory_blocks = build_mandatory_blocks(metadata, sections, cost_terms, policy_topics)
    terms = build_terms(metadata, cost_terms, policy_topics)

    full = {
        "meta": {
            "schema_version": "2.0",
            "output_version": OUTPUT_VERSION,
            "framework": FRAMEWORK_NAME,
        },
        "paper_info": {
            "title": metadata.title,
            "authors": metadata.authors,
            "year": metadata.year,
            "language": metadata.language,
            "field": metadata.field,
            "keywords": metadata.keywords,
        },
        "language_detect_result": {
            "detected_language": metadata.language,
            "confidence": metadata.language_confidence,
            "reason": metadata.language_reason,
        },
        "mandatory_blocks": mandatory_blocks,
        "one_sentence_summary": build_one_sentence_summary(metadata, mandatory_blocks),
        "general_interview": build_general_interview(mandatory_blocks),
        "english_interview": build_english_interview(mandatory_blocks, metadata, terms),
        "professional_written_exam": build_professional_written_exam(mandatory_blocks),
        "english_written_exam": build_english_written_exam(metadata, terms),
        "review_outline": build_review_outline(),
        "terms": terms,
        "low_priority": build_low_priority(metadata),
        "extra": build_extra(metadata),
    }
    slug, used_slug_fallback = slugify(metadata.title)
    return full, slug, used_title_fallback or used_slug_fallback, metadata, sections


def split_outputs(full: dict[str, Any]) -> dict[str, dict[str, Any]]:
    base = {
        "meta": full["meta"],
        "paper_info": full["paper_info"],
        "language_detect_result": full["language_detect_result"],
        "mandatory_blocks": full["mandatory_blocks"],
        "one_sentence_summary": full["one_sentence_summary"],
        "terms": full["terms"],
        "low_priority": full["low_priority"],
        "extra": full["extra"],
    }
    return {
        "general_interview.json": {
            **base,
            "general_interview": full["general_interview"],
            "review_outline": {"general_interview_outline": full["review_outline"]["general_interview_outline"]},
        },
        "english_interview.json": {
            **base,
            "english_interview": full["english_interview"],
            "review_outline": {"english_interview_outline": full["review_outline"]["english_interview_outline"]},
        },
        "professional_written_exam.json": {
            **base,
            "professional_written_exam": full["professional_written_exam"],
            "review_outline": {"professional_written_exam_outline": full["review_outline"]["professional_written_exam_outline"]},
        },
        "english_written_exam.json": {
            **base,
            "english_written_exam": full["english_written_exam"],
            "review_outline": {"english_written_exam_outline": full["review_outline"]["english_written_exam_outline"]},
        },
    }


def load_schema(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def validate_against_schema(data: Any, schema: Any, path: str = "$") -> None:
    if isinstance(schema, dict):
        if not isinstance(data, dict):
            raise ValueError(f"{path} should be an object")
        schema_keys = set(schema.keys())
        data_keys = set(data.keys())
        missing = schema_keys - data_keys
        extra = data_keys - schema_keys
        if missing:
            raise ValueError(f"{path} missing keys: {sorted(missing)}")
        if extra:
            raise ValueError(f"{path} has unexpected keys: {sorted(extra)}")
        for key, value in schema.items():
            validate_against_schema(data[key], value, f"{path}.{key}")
        return
    if isinstance(schema, list):
        if not isinstance(data, list):
            raise ValueError(f"{path} should be a list")
        if not schema:
            return
        exemplar = schema[0]
        for index, item in enumerate(data):
            validate_against_schema(item, exemplar, f"{path}[{index}]")
        return
    if isinstance(schema, str):
        if not isinstance(data, str):
            raise ValueError(f"{path} should be a string")
        return
    if isinstance(schema, bool):
        if not isinstance(data, bool):
            raise ValueError(f"{path} should be a boolean")
        return
    if isinstance(schema, (int, float)):
        if not isinstance(data, (int, float)):
            raise ValueError(f"{path} should be a number")
        return
    if schema is None and data is not None:
        raise ValueError(f"{path} should be null")


def write_json(path: Path, data: dict[str, Any], schema_path: Path) -> None:
    schema = load_schema(schema_path)
    validate_against_schema(data, schema)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    reloaded = json.loads(path.read_text(encoding="utf-8"))
    validate_against_schema(reloaded, schema)


def style_sheet(sheet, theme: str = "default") -> None:
    if theme == "memorize":
        header_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        title_fill = PatternFill(fill_type="solid", fgColor="9E2A2B")
        section_fill = PatternFill(fill_type="solid", fgColor="C6E0B4")
        title_font_color = "FFFFFF"
        header_font_color = "000000"
    elif theme == "print":
        header_fill = PatternFill(fill_type="solid", fgColor="E7E6E6")
        title_fill = PatternFill(fill_type="solid", fgColor="7F7F7F")
        section_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        title_font_color = "FFFFFF"
        header_font_color = "000000"
    else:
        header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
        title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        section_fill = PatternFill(fill_type="solid", fgColor="F4B183")
        title_font_color = "FFFFFF"
        header_font_color = "000000"

    for cell in sheet[1]:
        cell.font = Font(bold=True, color=title_font_color)
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if sheet.max_row >= 2:
        for cell in sheet[2]:
            cell.font = Font(bold=True, color=header_font_color)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 16), 56)

    if sheet.max_row >= 1:
        sheet.freeze_panes = "A3"

    if theme == "memorize":
        for row_idx in range(3, sheet.max_row + 1):
            if row_idx % 2 == 1:
                for cell in sheet[row_idx]:
                    if cell.fill.fill_type is None:
                        cell.fill = PatternFill(fill_type="solid", fgColor="FCFCF2")

    if theme == "print":
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_title_rows = "$1:$2"
        sheet.page_margins.left = 0.3
        sheet.page_margins.right = 0.3
        sheet.page_margins.top = 0.4
        sheet.page_margins.bottom = 0.4


def append_table(sheet, title: str, headers: list[str], rows: list[list[str]]) -> None:
    sheet.append([title])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.append([])


def create_excel_workbook(output_path: Path, full: dict[str, Any], run_report: dict[str, Any], theme: str = "default") -> None:
    workbook = Workbook()

    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Overview"])
    overview.append(["Field", "Value"])
    info = full["paper_info"]
    overview_rows = [
        ["Title", info["title"]],
        ["Authors", "；".join(info["authors"]) if info["authors"] else ""],
        ["Year", info["year"]],
        ["Language", info["language"]],
        ["Field", info["field"]],
        ["Keywords", "；".join(info["keywords"])],
        ["Summary", full["one_sentence_summary"]],
        ["Schema Version", full["meta"]["schema_version"]],
        ["Output Version", full["meta"]["output_version"]],
        ["Framework", full["meta"]["framework"]],
    ]
    for row in overview_rows:
        overview.append(row)
    style_sheet(overview, theme=theme)

    mandatory_sheet = workbook.create_sheet("Mandatory Blocks")
    append_table(
        mandatory_sheet,
        "Mandatory Blocks",
        ["Module", "Chinese Label", "Content"],
        [[module, module_label(module), full["mandatory_blocks"][module]] for module in MANDATORY_MODULES],
    )
    style_sheet(mandatory_sheet, theme=theme)

    general_sheet = workbook.create_sheet("General Interview")
    append_table(
        general_sheet,
        "General Interview",
        ["Module", "Question", "Answer", "Why This Matters"],
        [[item["module"], item["question"], item["answer"], item["why_this_matters"]] for item in full["general_interview"]],
    )
    style_sheet(general_sheet, theme=theme)

    english_interview_sheet = workbook.create_sheet("English Interview")
    append_table(
        english_interview_sheet,
        "English Interview",
        ["Module", "Question", "Answer", "Terminology Notes"],
        [[item["module"], item["question_en"], item["answer_en"], item["terminology_notes"]] for item in full["english_interview"]],
    )
    style_sheet(english_interview_sheet, theme=theme)

    professional_sheet = workbook.create_sheet("Professional Written")
    append_table(
        professional_sheet,
        "Professional Written",
        ["Module", "Question", "Answer", "Answer Type"],
        [[item["module"], item["question"], item["answer"], item["answer_type"]] for item in full["professional_written_exam"]],
    )
    style_sheet(professional_sheet, theme=theme)

    english_written_sheet = workbook.create_sheet("English Written")
    append_table(
        english_written_sheet,
        "English Written",
        ["Module", "Question", "Answer", "Answer Type"],
        [[item["module"], item["question_en"], item["answer_en"], item["answer_type"]] for item in full["english_written_exam"]],
    )
    style_sheet(english_written_sheet, theme=theme)

    terms_sheet = workbook.create_sheet("Terms")
    append_table(
        terms_sheet,
        "Terms",
        ["Chinese", "English", "Explanation"],
        [[item["chinese"], item["english"], item["explanation"]] for item in full["terms"]],
    )
    style_sheet(terms_sheet, theme=theme)

    report_sheet = workbook.create_sheet("Run Report")
    report_sheet.append(["Run Report"])
    report_sheet.append(["Field", "Value"])
    for key, value in run_report.items():
        report_sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value])
    style_sheet(report_sheet, theme=theme)

    workbook.save(output_path)
    reloaded = load_workbook(output_path)
    expected_sheets = [
        "Overview",
        "Mandatory Blocks",
        "General Interview",
        "English Interview",
        "Professional Written",
        "English Written",
        "Terms",
        "Run Report",
    ]
    if reloaded.sheetnames != expected_sheets:
        raise ValueError(f"Excel workbook sheets mismatch: {reloaded.sheetnames}")


def build_run_report(source: SourceData, metadata: PaperMetadata, sections: PaperSections, used_fallback_slug: bool) -> dict[str, Any]:
    return {
        "input_type": source.input_type,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "language": metadata.language,
        "paper_type": metadata.paper_type,
        "used_fallback_slug": used_fallback_slug,
        "abstract_length": len(sections.abstract),
        "source_path": source.source_path,
        "output_version": OUTPUT_VERSION,
    }


def main() -> int:
    args = parse_args()
    input_path = args.input_file.resolve() if args.input_file else None
    try:
        source = load_source(input_path, args.input_text)
        full, auto_slug, used_fallback_slug, metadata, sections = build_full_output(source, forced_language=args.language)
        split_files = split_outputs(full)

        if args.stdout_json:
            print(json.dumps(full, ensure_ascii=False, indent=2))
            return 0

        slug = args.slug or auto_slug
        output_root = args.output_dir.resolve() if args.output_dir else Path.cwd() / "output" / "economics-retest-paper-splitter"
        output_dir = output_root / slug
        output_dir.mkdir(parents=True, exist_ok=True)
        run_report = build_run_report(source, metadata, sections, used_fallback_slug)

        write_json(output_dir / "full.json", full, FULL_SCHEMA_PATH)
        schema_map = {
            "general_interview.json": GENERAL_SCHEMA_PATH,
            "english_interview.json": ENGLISH_INTERVIEW_SCHEMA_PATH,
            "professional_written_exam.json": PROFESSIONAL_WRITTEN_SCHEMA_PATH,
            "english_written_exam.json": ENGLISH_WRITTEN_SCHEMA_PATH,
        }
        for filename, payload in split_files.items():
            write_json(output_dir / filename, payload, schema_map[filename])

        (output_dir / "run-report.json").write_text(
            json.dumps(run_report, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        create_excel_workbook(output_dir / "retest_pack.xlsx", full, run_report, theme="default")
        create_excel_workbook(output_dir / "retest_pack_memorize.xlsx", full, run_report, theme="memorize")
        create_excel_workbook(output_dir / "retest_pack_print.xlsx", full, run_report, theme="print")

        for filename in PRINTED_OUTPUTS:
            print(output_dir / filename)
        return 0
    except UserFacingError as exc:
        print(str(exc))
        return exc.code
    except ValueError as exc:
        print(str(exc))
        return EXIT_SCHEMA_INVALID


if __name__ == "__main__":
    sys.exit(main())
