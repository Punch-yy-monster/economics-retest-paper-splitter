#!/usr/bin/env python3
"""Run regression checks against standardized four-channel fixtures."""

from __future__ import annotations

import filecmp
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "generate_retest_json.py"
EXAMPLES = ROOT / "examples"
SUCCESS_FIXTURES = ("english-digital-economics", "chinese-digital-economics")
FAILURE_FIXTURES = ("failure-title-only-english", "failure-title-only-chinese")
MANDATORY_MODULES = [
    "research_background",
    "research_question",
    "core_conclusion",
    "mechanism_analysis",
    "policy_implication",
]
JSON_FILENAMES = (
    "full.json",
    "general_interview.json",
    "english_interview.json",
    "professional_written_exam.json",
    "english_written_exam.json",
)
EXCEL_FILENAMES = (
    "retest_pack.xlsx",
    "retest_pack_memorize.xlsx",
    "retest_pack_print.xlsx",
)


def run_command(args: list[str], cwd: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(args, cwd=cwd, capture_output=True, text=True)


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def assert_snapshot(generated: Path, expected: Path) -> None:
    if not filecmp.cmp(generated, expected, shallow=False):
        raise RuntimeError(f"snapshot mismatch: {generated.name}")


def assert_channel_modules(items: list[dict], question_key: str, answer_key: str) -> None:
    modules = [item["module"] for item in items]
    if modules != MANDATORY_MODULES:
        raise RuntimeError(f"channel module order mismatch: {modules!r}")
    for item in items:
        if not item[question_key] or not item[answer_key]:
            raise RuntimeError(f"channel item is missing content: {item!r}")


def assert_common_structure(full: dict, split_payloads: dict[str, dict]) -> None:
    meta = full.get("meta", {})
    if meta.get("schema_version") != "2.0":
        raise RuntimeError("full.json is missing meta.schema_version=2.0")
    if meta.get("output_version") != "2.0.0":
        raise RuntimeError("full.json is missing meta.output_version=2.0.0")
    if meta.get("framework") != "five-mandatory-modules-four-channels":
        raise RuntimeError("full.json framework marker is incorrect")

    mandatory_blocks = full.get("mandatory_blocks", {})
    if list(mandatory_blocks.keys()) != MANDATORY_MODULES:
        raise RuntimeError("mandatory_blocks should keep the fixed five-module order")
    for module in MANDATORY_MODULES:
        if not mandatory_blocks[module]:
            raise RuntimeError(f"mandatory_blocks[{module}] should not be empty")

    assert_channel_modules(full["general_interview"], "question", "answer")
    assert_channel_modules(full["english_interview"], "question_en", "answer_en")
    assert_channel_modules(full["professional_written_exam"], "question", "answer")
    assert_channel_modules(full["english_written_exam"], "question_en", "answer_en")

    split_expectations = {
        "general_interview.json": "general_interview",
        "english_interview.json": "english_interview",
        "professional_written_exam.json": "professional_written_exam",
        "english_written_exam.json": "english_written_exam",
    }
    for filename, channel_key in split_expectations.items():
        payload = split_payloads[filename]
        if payload["mandatory_blocks"] != mandatory_blocks:
            raise RuntimeError(f"{filename} should reuse full.json mandatory_blocks")
        if channel_key not in payload:
            raise RuntimeError(f"{filename} is missing {channel_key}")


def assert_example_expectations(example_name: str, full: dict) -> None:
    paper_info = full["paper_info"]
    if example_name == "english-digital-economics":
        if paper_info["title"] != "Digital Economics":
            raise RuntimeError("english example title extraction failed")
        if paper_info["language"] != "英文文献":
            raise RuntimeError("english example language detection failed")
        if not paper_info["authors"]:
            raise RuntimeError("english example author extraction failed")
        if "成本结构" not in full["mandatory_blocks"]["core_conclusion"]:
            raise RuntimeError("english review example should emphasize cost structure in core conclusion")
    elif example_name == "chinese-digital-economics":
        if paper_info["title"] != "数据要素流通、平台治理与区域创新":
            raise RuntimeError("chinese example title extraction failed")
        if paper_info["language"] != "中文文献":
            raise RuntimeError("chinese example language detection failed")
        if not {"数据要素流通", "平台治理", "区域创新"}.issubset(set(paper_info["keywords"])):
            raise RuntimeError("chinese example keyword fallback failed")
        if "数字治理" not in full["mandatory_blocks"]["policy_implication"] and "制度" not in full["mandatory_blocks"]["policy_implication"]:
            raise RuntimeError("chinese empirical example should surface policy implications")


def run_success_fixture(example_name: str) -> None:
    example_dir = EXAMPLES / example_name
    input_file = example_dir / "input.txt"
    expected_dir = example_dir / "expected"

    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        output_root = tmpdir / "generated"
        result = run_command(
            [sys.executable, str(SCRIPT), "--input-file", str(input_file), "--output-dir", str(output_root)],
            cwd=tmpdir,
        )
        if result.returncode != 0:
            raise RuntimeError(f"{example_name}: command failed: {result.stdout or result.stderr}")

        paths = [Path(line.strip()) for line in result.stdout.splitlines() if line.strip()]
        if len(paths) != 8:
            raise RuntimeError(f"{example_name}: expected 8 printed output paths, got {paths!r}")

        generated_dir = paths[0].parent
        report_path = generated_dir / "run-report.json"
        if not report_path.exists():
            raise RuntimeError(f"{example_name}: run-report.json not generated")

        full = load_json(generated_dir / "full.json")
        split_payloads = {filename: load_json(generated_dir / filename) for filename in JSON_FILENAMES if filename != "full.json"}
        report = load_json(report_path)

        assert_common_structure(full, split_payloads)
        assert_example_expectations(example_name, full)

        if report["input_type"] not in {"text", "txt"}:
            raise RuntimeError(f"{example_name}: unexpected input_type {report['input_type']}")
        if report["abstract_length"] <= 0:
            raise RuntimeError(f"{example_name}: abstract_length should be positive")
        if report["output_version"] != "2.0.0":
            raise RuntimeError(f"{example_name}: unexpected output version {report['output_version']!r}")

        expected_sheets = [
            "Overview",
            "Mandatory Blocks",
            "General Interview",
            "English Interview",
            "Professional Written",
            "English Written",
            "Terms",
            "Run Report",
        ]
        for workbook_name in EXCEL_FILENAMES:
            workbook = load_workbook(generated_dir / workbook_name)
            if workbook.sheetnames != expected_sheets:
                raise RuntimeError(f"{example_name}: unexpected Excel sheet names {workbook.sheetnames!r}")

        for filename in JSON_FILENAMES:
            assert_snapshot(generated_dir / filename, expected_dir / filename)


def run_failure_fixture(example_name: str) -> None:
    example_dir = EXAMPLES / example_name
    input_file = example_dir / "input.txt"
    expected_error = (example_dir / "expected_error.txt").read_text(encoding="utf-8").strip()
    result = run_command([sys.executable, str(SCRIPT), "--input-file", str(input_file)], cwd=ROOT)
    if result.returncode != 1:
        raise RuntimeError(f"{example_name}: expected exit code 1, got {result.returncode}")
    if result.stdout.strip() != expected_error:
        raise RuntimeError(f"{example_name}: expected error message {expected_error!r}, got {result.stdout.strip()!r}")


def run_cli_checks() -> None:
    chinese_input = (EXAMPLES / "chinese-digital-economics" / "input.txt").read_text(encoding="utf-8")
    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)

        stdout_result = run_command(
            [sys.executable, str(SCRIPT), "--input-text", chinese_input, "--stdout-json", "--language", "中文文献"],
            cwd=tmpdir,
        )
        if stdout_result.returncode != 0:
            raise RuntimeError(f"stdout-json check failed: {stdout_result.stdout or stdout_result.stderr}")
        stdout_json = json.loads(stdout_result.stdout)
        if list(stdout_json["mandatory_blocks"].keys()) != MANDATORY_MODULES:
            raise RuntimeError("--stdout-json should keep fixed mandatory blocks")
        if (tmpdir / "output").exists():
            raise RuntimeError("--stdout-json should not create output directory")

        slug_root = tmpdir / "cli-output"
        slug_result = run_command(
            [
                sys.executable,
                str(SCRIPT),
                "--input-text",
                chinese_input,
                "--language",
                "中文文献",
                "--slug",
                "custom-slug",
                "--output-dir",
                str(slug_root),
            ],
            cwd=tmpdir,
        )
        if slug_result.returncode != 0:
            raise RuntimeError(f"--slug check failed: {slug_result.stdout or slug_result.stderr}")
        if not (slug_root / "custom-slug" / "full.json").exists():
            raise RuntimeError("--slug should control the output directory")

        mixed_text = (
            "Title: Platform Governance and 数据要素流通\n\n"
            "Abstract: This paper studies how platform governance affects data circulation and regional innovation. "
            "It argues that digital rules shape incentives, information flows, and organizational choices in platform markets. "
            "文章同时讨论平台治理、数据流通与区域创新之间的关系，并强调制度环境、信息不对称缓解和资源配置效率改善的重要性。"
        )
        mixed_result = run_command([sys.executable, str(SCRIPT), "--input-text", mixed_text, "--stdout-json"], cwd=tmpdir)
        if mixed_result.returncode != 0:
            raise RuntimeError(f"mixed-language check failed: {mixed_result.stdout or mixed_result.stderr}")
        mixed_json = json.loads(mixed_result.stdout)
        if mixed_json["language_detect_result"]["detected_language"] != "中英混合文献":
            raise RuntimeError("mixed-language text should be classified as 中英混合文献")


def main() -> int:
    for example_name in SUCCESS_FIXTURES:
        run_success_fixture(example_name)
        print(f"{example_name}: ok")

    for example_name in FAILURE_FIXTURES:
        run_failure_fixture(example_name)
        print(f"{example_name}: ok")

    run_cli_checks()
    print("cli-checks: ok")
    return 0


if __name__ == "__main__":
    sys.exit(main())
